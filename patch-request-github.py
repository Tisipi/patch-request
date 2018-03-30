import yaml
import openpyxl


class DeviceDefinitionsFile:
    """
    The Device Definitions are imported from a file in YAML format.
    This file contains a list of devices with its attributes. A device entry looks like this:
        aDeviceName:
            model: aModel
            device_id: aNumber
            room: aRoom
            rack: aRack
            location: aLocation
    From this file a device dictionary is built using the method build_device_dictionary
    """

    def __init__(self, file):
        self.file = file

    def print(self):
        with open(self.file, 'r') as f:
            print('Contents of Device Definitions File {}:'.format(self.file))
            for line in f:
                print(line.rstrip())

    def build_device_dictionary(self):
        print('Building device dictionary from file {} '.format(self.file))
        with open(self.file, 'r') as f:
            device_dict = yaml.load(f)
        return device_dict


class ConnectivityRequest:
    """
    A Connectivity Request is a list of requests to create or remove (patch) cables between devices.
    A connectivity request has this format:
    Create <source device> <source port> - <destination device> <destination port>
    Remove <source device> <source port> - <destination device> <destination port>
    """

    def __init__(self, file):
        self.file = file

    def print(self):
        with open(self.file, 'r') as f:
            print('Contents of Connectivity Request {}:'.format(self.file))
            for line in f:
                print(line.rstrip())

    def list_requests(self):
        print('Retrieving requests from connectivity request {}'.format(self.file))
        requests = []
        with open(self.file, 'r') as f:
            for line in f:
                patch_request = line.split()
                # E.g. Create/Remove DC-NLUT-034 Gi1/15 - PE-NLUT-001 Gi0/2/0/30
                action = patch_request[0]
                src_dev = patch_request[1]
                src_int = patch_request[2]
                dest_dev = patch_request[4]
                dest_int = patch_request[5]
                requests.append(Request(action, Connection(src_dev, src_int, dest_dev, dest_int)))
        return requests


class Request:
    def __init__(self, action, connection):
        if action not in ['Create', 'Remove']:
            print('Request {} is not allowed. Please use Create and Remove'.format(action))
        self.action = action
        self.connection = connection

    def get_action(self):
        return self.action

    def source_device(self):
        return self.connection.source_device()

    def source_port(self):
        return self.connection.source_port()

    def destination_device(self):
        return self.connection.destination_device()

    def destination_port(self):
        return self.connection.destination_port()

    def print(self):
        print('Request to {} {}'.format(self.action, self.connection.asString()))


class Connection:
    def __init__(self, src_dev, src_int, dest_dev, dest_int):
        self.src_dev = src_dev
        self.src_int = src_int
        self.dest_dev = dest_dev
        self.dest_int = dest_int

    def source_device(self):
        return self.src_dev

    def source_port(self):
        return self.src_int

    def destination_device(self):
        return self.dest_dev

    def destination_port(self):
        return self.dest_int

    def asString(self):
        return self.src_dev + ' ' + self.src_int + ' ' + self.dest_dev + ' ' + self.dest_int


class PatchRequest:
    """
    A Patch Request is an excel file that is sent to the cabling team.
    This team patches the cables between the devices.
    A patch request is a connectivity request enhanced with information like datacenter room, rack, etc.
    """

    def __init__(self, file):
        self.file = file

    def build(self, list_of_requests, device_dict):
        # os.chdir('C:\\Logs')
        print('Generating patch request {}'.format(self.file))
        patch_request_wb = openpyxl.Workbook()
        # Remove the default created sheet called 'Sheet'.
        patch_request_wb.remove(patch_request_wb['Sheet'])
        # Create sheet in workbook.
        patch_request_wb.create_sheet(title='Cabling Request')
        sheet = patch_request_wb['Cabling Request']
        self._build_sheet(sheet, list_of_requests, device_dict)
        patch_request_wb.save(self.file)

    def _build_sheet(self, sheet, list_of_requests, device_dict):
        start_row = 1
        self._build_sheet_header(sheet, start_row)
        self._build_sheet_rows(sheet, start_row + 1, list_of_requests, device_dict)

    def _build_sheet_header(self, sheet, start_row):
        header = ["Action", "Device", "Model", "Device ID", "Port", "Location", "Room", "Rack", "Cable Type", "Device", "Model", "Device ID", "Port", "Room", "Rack"]
        for i in range(len(header)):
            column_number = i + 1
            sheet.cell(row=start_row, column=column_number).value = header[i]

    def _build_sheet_rows(self, sheet, start_row, list_of_requests, device_dict):
        for i in range(len(list_of_requests)):
            row_number = start_row + i
            request = list_of_requests[i]
            source_device = request.source_device()
            destination_device = request.destination_device()
            sheet.cell(row=row_number, column=1).value = request.get_action()
            sheet.cell(row=row_number, column=2).value = source_device
            sheet.cell(row=row_number, column=3).value = device_dict[source_device]['model']
            sheet.cell(row=row_number, column=4).value = device_dict[source_device]['device_id']
            sheet.cell(row=row_number, column=5).value = request.source_port()
            sheet.cell(row=row_number, column=6).value = device_dict[source_device]['location']
            sheet.cell(row=row_number, column=7).value = device_dict[source_device]['room']
            sheet.cell(row=row_number, column=8).value = device_dict[source_device]['rack']
            sheet.cell(row=row_number, column=9).value = 'RJ45'
            sheet.cell(row=row_number, column=10).value = destination_device
            sheet.cell(row=row_number, column=11).value = device_dict[destination_device]['model']
            sheet.cell(row=row_number, column=12).value = device_dict[destination_device]['device_id']
            sheet.cell(row=row_number, column=13).value = request.destination_port()
            sheet.cell(row=row_number, column=14).value = device_dict[destination_device]['room']
            sheet.cell(row=row_number, column=15).value = device_dict[destination_device]['rack']


# Location of files
path = 'C:\\Templates\\'
device_properties_file = path + 'deviceInfo.yaml'
connectivity_request_file = path + 'connectivity request.txt'
patch_excel_file = path + 'patch request.xlsx'

# Main program
device_dict = DeviceDefinitionsFile(device_properties_file).build_device_dictionary()
list_of_requests = ConnectivityRequest(connectivity_request_file).list_requests()
PatchRequest(patch_excel_file).build(list_of_requests, device_dict)
